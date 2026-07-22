# report_excel.py
# Escritura y formateo de los resultados del análisis estructural en una hoja de cálculo.
#
# Uso desde linea de comandos:
# N/A. Se importa como módulo: from report_excel import main
#
# Flujo:
# 1. Carga el libro de trabajo (Workbook) base usando openpyxl.
# 2. Escribe los datos de entrada (propiedades, apoyos, cargas).
# 3. Escribe los resultados de salida (Mmax, módulo de sección requerido, reacciones).
# 4. Aplica estilos corporativos (colores, fuentes, bordes) a las celdas.
# 5. Escribe un registro (log) de ejecución y guarda el archivo .xlsx actualizado.


import os
import traceback
from datetime import datetime

import numpy as np
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# ── Constantes de diseño ─────────────────────────────────────────────────────
TYPESLOADS = [
    "Point load",
    "Uniformly distributed",
    "Triangular distributed 1",
    "Triangular distributed 2",
    "Moment load",
]

TYPES_SUPPORT = {"Roller": 1, "Pinned": 1, "Fixed": 2}


# Colores corporativos (ARGB sin '#')
C_BLUE_DARK  = "FF2E5B8A"
C_BLUE_LIGHT = "FFD6E4F0"
C_WHITE      = "FFFFFFFF"
C_GRAY_LIGHT = "FFF1EFE8"
C_GREEN_OK   = "FF1A7A3A"
C_RED_ERR    = "FFB22222"
C_AMBER      = "FFBA7517"

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS DE FORMATO
# ─────────────────────────────────────────────────────────────────────────────

def _header_style(cell, bg=C_BLUE_DARK, fg=C_WHITE, bold=True, size=11):
    cell.font      = Font(bold=bold, color=fg, size=size, name="Courier New")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)

def _data_style(cell, bold=False, color="FF2C2C2A", bg=None, align="center"):
    cell.font      = Font(bold=bold, color=color, size=10, name="Courier New")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)

def _border_all(cell, style="thin"):
    s = Side(style=style)
    cell.border = Border(left=s, right=s, top=s, bottom=s)

def _section_title(ws, row, col, text, width=6):
    """Escribe un título de sección que abarca 'width' columnas."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(bold=True, color=C_WHITE, size=12, name="Courier New")
    cell.fill      = PatternFill("solid", fgColor=C_BLUE_DARK)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                                indent=1)
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + width - 1)
    ws.row_dimensions[row].height = 22

def _write_table(ws, start_row, start_col, headers, rows, alt_bg=C_BLUE_LIGHT):
    """Escribe una tabla booktabs-style con cabecera azul y filas alternas."""
    # Cabecera
    for c, h in enumerate(headers, start=start_col):
        cell = ws.cell(row=start_row, column=c, value=h)
        _header_style(cell)
        _border_all(cell)
    ws.row_dimensions[start_row].height = 18

    # Filas
    for r_idx, row in enumerate(rows):
        bg = alt_bg if r_idx % 2 == 1 else C_WHITE
        for c_idx, val in enumerate(row, start=start_col):
            cell = ws.cell(row=start_row + 1 + r_idx, column=c_idx, value=val)
            _data_style(cell, bg=bg)
            _border_all(cell, style="hair")
        ws.row_dimensions[start_row + 1 + r_idx].height = 16

    return start_row + 1 + len(rows)   # devuelve la fila siguiente a la tabla

# ─────────────────────────────────────────────────────────────────────────────
# 4. ESCRITURA DE RESULTADOS EN EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def write_results(wb, r, f, l, Sy, FS, r_sol, Mmax, s, img_dcl, img_vm):
    """Sobreescribe la hoja 'Resultados' con tablas e imágenes."""

    # Eliminar hoja anterior y crear nueva
    if "Resultados" in wb.sheetnames:
        del wb["Resultados"]
    ws = wb.create_sheet("Resultados")

    # Mover la hoja a la posición 1 (después de Inputs)
    idx_inputs = wb.sheetnames.index("Inputs")
    wb.move_sheet("Resultados", offset=idx_inputs + 1 - wb.sheetnames.index("Resultados"))

    # ── Anchos de columna ────────────────────────────────────────────────────
    col_widths = [2, 18, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Título principal ─────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    cell = ws.cell(row=1, column=2,
                   value="PyBeams  —  Resultados del Analisis Estructural")
    cell.font      = Font(bold=True, color=C_WHITE, size=14, name="Courier New")
    cell.fill      = PatternFill("solid", fgColor=C_BLUE_DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B1:H1")

    ts = datetime.now().strftime("%d/%m/%Y  %H:%M:%S")
    ws.cell(row=2, column=2, value=f"Generado: {ts}").font = Font(
        italic=True, color="FF888780", size=9, name="Courier New"
    )
    ws.merge_cells("B2:H2")

    cur_row = 4

    # ── § 1  PARAMETROS DE ENTRADA ───────────────────────────────────────────
    _section_title(ws, cur_row, 2, "1.  Parametros de entrada", width=7)
    cur_row += 1

    sigmaAllow = Sy / FS
    param_rows = [
        ["Longitud de la viga (l)",           f"{l:.3f}",          "m"],
        ["Esfuerzo de fluencia (Sy)",         f"{Sy:.1f}",         "MPa"],
        ["Factor de Seguridad (FS)",          f"{FS:.2f}",         "---"],
        ["Esfuerzo admisible (sigma_adm)",    f"{sigmaAllow:.2f}", "MPa"],
    ]
    cur_row = _write_table(ws, cur_row, 2,
                           ["Parametro", "Valor", "Unidad"], param_rows) + 1

    # ── § 2  APOYOS ──────────────────────────────────────────────────────────
    _section_title(ws, cur_row, 2, "2.  Configuracion de apoyos", width=7)
    cur_row += 1

    dof_desc = {1: "1 reaccion (Fy)", 2: "2 reacciones (Fy, M)"}
    apoyo_rows = [
        [k,
         f"{v['location']:.3f}",
         v["type"],
         dof_desc.get(v["dof"], "---")]
        for k, v in r.items()
    ]
    cur_row = _write_table(ws, cur_row, 2,
                           ["ID", "Ubicacion [m]", "Tipo", "Restricciones"],
                           apoyo_rows) + 1

    # ── § 3  CARGAS APLICADAS ────────────────────────────────────────────────
    _section_title(ws, cur_row, 2, "3.  Cargas aplicadas", width=7)
    cur_row += 1

    unit_map = {
        TYPESLOADS[0]: "kN",   TYPESLOADS[1]: "kN/m",
        TYPESLOADS[2]: "kN/m", TYPESLOADS[3]: "kN/m",
        TYPESLOADS[4]: "kN.m",
    }
    carga_rows = [
        [k,
         f"{v['value']:.3f}",
         f"{v['a']:.3f}",
         f"{v.get('a1', v['a']):.3f}",
         v["type"],
         unit_map.get(v["type"], "---")]
        for k, v in f.items()
    ]
    cur_row = _write_table(ws, cur_row, 2,
                           ["ID", "Magnitud", "a [m]", "a1 [m]",
                            "Tipo", "Unidad"],
                           carga_rows) + 1

    # ── § 4  REACCIONES ──────────────────────────────────────────────────────
    _section_title(ws, cur_row, 2, "4.  Reacciones en los apoyos", width=7)
    cur_row += 1

    react_rows = []
    for name, data in r_sol.items():
        val  = data["value"]
        pos  = data["a"]
        tipo = data["type"]
        unid = "kN" if tipo == TYPESLOADS[0] else "kN.m"
        sent = "positivo" if val >= 0 else "negativo"
        react_rows.append([name, f"{pos:.3f}", f"{val:.4f}", unid, sent])

    cur_row = _write_table(ws, cur_row, 2,
                           ["Reaccion", "Posicion [m]", "Valor",
                            "Unidad", "Sentido"],
                           react_rows) + 1

    # ── § 5  VERIFICACION DE EQUILIBRIO ─────────────────────────────────────
    _section_title(ws, cur_row, 2, "5.  Verificacion de equilibrio", width=7)
    cur_row += 1

    sum_Fy = sum_M0 = 0.0
    eq_rows = []

    for key, load in f.items():
        val  = load["value"]
        a    = load["a"]
        a1   = load.get("a1", a)
        ltyp = load["type"]
        if ltyp == TYPESLOADS[0]:
            Fy = val;  M0 = val * a;                   desc = f"{key} puntual"
        elif ltyp == TYPESLOADS[1]:
            lg = a1 - a; Fy = val * lg; M0 = Fy*(a+lg/2); desc = f"{key} dist.unif."
        elif ltyp == TYPESLOADS[2]:
            lg = a1-a; Fy=val*lg/2; M0=Fy*(a+lg*2/3); desc = f"{key} tri.crec."
        elif ltyp == TYPESLOADS[3]:
            lg = abs(a-a1); Fy=val*lg/2; M0=Fy*(a+lg/3); desc = f"{key} tri.dec."
        elif ltyp == TYPESLOADS[4]:
            Fy = 0.0; M0 = val;                        desc = f"{key} momento"
        else:
            continue
        sum_Fy += Fy;  sum_M0 += M0
        eq_rows.append([desc, f"{Fy:.4f}", f"{M0:.4f}"])

    for name, data in r_sol.items():
        val = data["value"]; pos = data["a"]; tipo = data["type"]
        if tipo == TYPESLOADS[0]:
            Fy = val; M0 = val*pos; desc = f"R_{name}"
        else:
            Fy = 0.0; M0 = val;    desc = f"M_{name}"
        sum_Fy += Fy; sum_M0 += M0
        eq_rows.append([desc, f"{Fy:.4f}", f"{M0:.4f}"])

    cur_row = _write_table(ws, cur_row, 2,
                           ["Elemento", "Fy [kN]", "M_O [kN.m]"],
                           eq_rows)

    # Resultado de equilibrio
    tol = 1e-4
    ok_Fy = abs(sum_Fy) < tol
    ok_M  = abs(sum_M0) < tol

    for col_off, label, val_num, ok in [
        (0, "Sum Fy [kN]",   sum_Fy, ok_Fy),
        (2, "Sum M_O [kN.m]", sum_M0, ok_M),
    ]:
        lbl_cell = ws.cell(row=cur_row, column=2 + col_off, value=label)
        _data_style(lbl_cell, bold=True)
        val_cell = ws.cell(row=cur_row, column=3 + col_off, value=round(val_num, 6))
        _data_style(val_cell)
        sta_cell = ws.cell(row=cur_row, column=4 + col_off,
                           value="OK" if ok else "ERROR")
        sta_cell.font = Font(bold=True, color=C_GREEN_OK if ok else C_RED_ERR,
                              size=10, name="Courier New")
        sta_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[cur_row].height = 18
    cur_row += 2

    # ── § 6  VALORES CRITICOS ────────────────────────────────────────────────
    _section_title(ws, cur_row, 2, "6.  Valores criticos", width=7)
    cur_row += 1

    crit_rows = [
        ["Momento flector maximo  |M|_max", f"{Mmax:.4f}", "kN.m"],
        ["Modulo de seccion requerido  s",  f"{s:.2f}",    "cm3"],
    ]
    cur_row = _write_table(ws, cur_row, 2,
                           ["Magnitud", "Valor", "Unidad"], crit_rows) + 2

    # ── § 7  DIAGRAMAS ───────────────────────────────────────────────────────
    _section_title(ws, cur_row, 2, "7.  Diagrama de cuerpo libre", width=7)
    cur_row += 1

    if img_dcl and os.path.exists(img_dcl):
        img = XLImage(img_dcl)
        img.width  = 620
        img.height = 380
        anchor = f"B{cur_row}"
        ws.add_image(img, anchor)
        cur_row += 14      # reservar espacio para la imagen (aprox. 14 filas)
    else:
        ws.cell(row=cur_row, column=2,
                value="[Imagen DCL no encontrada]").font = Font(
            italic=True, color=C_RED_ERR, name="Courier New")
        cur_row += 1

    cur_row += 1
    _section_title(ws, cur_row, 2,
                   "8.  Diagramas de Fuerza Cortante y Momento Flector",
                   width=7)
    cur_row += 1

    if img_vm and os.path.exists(img_vm):
        img2 = XLImage(img_vm)
        img2.width  = 620
        img2.height = 380
        ws.add_image(img2, f"B{cur_row}")
        cur_row += 24
    else:
        ws.cell(row=cur_row, column=2,
                value="[Imagen V-M no encontrada]").font = Font(
            italic=True, color=C_RED_ERR, name="Courier New")
        cur_row += 1

    # Congelar paneles bajo el título
    ws.freeze_panes = "B3"

# ─────────────────────────────────────────────────────────────────────────────
# 5. ESCRITURA DEL LOG
# ─────────────────────────────────────────────────────────────────────────────

def write_log(wb, status, message="", elapsed=None):
    """Añade una fila al historial en la hoja 'Log'."""
    if "Log" not in wb.sheetnames:
        ws = wb.create_sheet("Log")
        ws.append(["Timestamp", "Estado", "Mensaje", "Tiempo (s)"])
        for col, w in [(1,20),(2,12),(3,60),(4,12)]:
            ws.column_dimensions[get_column_letter(col)].width = w
        _header_style(ws.cell(row=1, column=1))
        _header_style(ws.cell(row=1, column=2))
        _header_style(ws.cell(row=1, column=3))
        _header_style(ws.cell(row=1, column=4))
        ws.freeze_panes = "A2"
    else:
        ws = wb["Log"]

    ts      = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    elapsed_str = f"{elapsed:.1f}" if elapsed is not None else "---"
    ws.append([ts, status, message, elapsed_str])

    last_row = ws.max_row
    color    = C_GREEN_OK if status == "OK" else C_RED_ERR
    ws.cell(row=last_row, column=2).font = Font(bold=True, color=color,
                                                 name="Courier New", size=10)
    ws.row_dimensions[last_row].height = 16


# ─────────────────────────────────────────────────────────────────────────────
# 6. ENTRADA PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def main(xlsx_path, r, f, l, Sy, FS, r_sol, Mmax, s, img_dcl, img_vm):
    t_start = datetime.now()
    print(f"[PyBeams] Abriendo: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path)

    try:
        # Paso 4: escribir resultados
        write_results(wb, r, f, l, Sy, FS, r_sol, Mmax, s, img_dcl, img_vm)
        print("[PyBeams] Hoja 'Resultados' actualizada")

        elapsed = (datetime.now() - t_start).total_seconds()
        write_log(wb, "OK",
                  f"l={l}m | Mmax={Mmax:.2f} kN.m | s={s:.1f} cm3",
                  elapsed)

    except Exception as exc:
        msg = traceback.format_exc()
        print(f"[PyBeams] ERROR:\n{msg}")
        elapsed = (datetime.now() - t_start).total_seconds()
        write_log(wb, "ERROR", str(exc)[:120], elapsed)

    finally:
        wb.save(xlsx_path)
        print(f"[PyBeams] Libro guardado: {xlsx_path}")



